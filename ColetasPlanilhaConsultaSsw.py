from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import time
import os  # Para abrir o Excel no final


# Função para ler coletas do Excel (apenas visíveis)
def ler_coletas_excel(arquivo_excel):
    wb = load_workbook(arquivo_excel, data_only=True)
    ws = wb['ontem']

    coletas = []
    for row in reversed(list(ws.iter_rows(min_row=2))):  # pula cabeçalho
        row_number = row[0].row
        if not ws.row_dimensions[row_number].hidden:
            valor = row[0].value
            try:
                numero = int(valor)
                coletas.append(str(numero))
            except:
                pass
    return coletas


# Caminho do Excel
arquivo_excel = r"C:\Users\flavi\OneDrive\Weslley\AcompanhamentoWeslley.xlsx"
coletas = ler_coletas_excel(arquivo_excel)

# --- LOGIN E CONSULTA NO SISTEMA ---
LOGIN_URL = "http://sistema.ssw.inf.br"
DOM = 'tki'
CPF = '11744601640'
USUARIO = 'weslley'
SENHA = 'W@tki111'

XPATH_DOM = '/html/body/form/input[1]'
XPATH_CPF = '/html/body/form/input[2]'
XPATH_USUARIO = '/html/body/form/input[3]'
XPATH_SENHA = '/html/body/form/input[4]'
XPATH_BOTAO_LOGIN = '/html/body/form/a'

# Iniciar navegador (com webdriver-manager)
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)
driver.maximize_window()
driver.get(LOGIN_URL)

try:
    # Login
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, XPATH_DOM)))
    driver.find_element(By.XPATH, XPATH_DOM).send_keys(DOM)
    driver.find_element(By.XPATH, XPATH_CPF).send_keys(CPF)
    driver.find_element(By.XPATH, XPATH_USUARIO).send_keys(USUARIO)
    driver.find_element(By.XPATH, XPATH_SENHA).send_keys(SENHA)
    driver.find_element(By.XPATH, XPATH_BOTAO_LOGIN).click()

    # Aguardar a nova página e digitar 103
    WebDriverWait(driver, 10).until(EC.staleness_of(driver.find_element(By.XPATH, XPATH_CPF)))
    campo_103 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, '3')))
    campo_103.send_keys("103")
    print("✅ Digitado 103 para abrir nova guia")

    time.sleep(3)  # Aguarda nova guia abrir
    driver.switch_to.window(driver.window_handles[-1])  # Foca na nova guia

    #Ajustando contador
    qtdcoletas = len(coletas)
    contador = 0

    for coleta in coletas:
        campo_coleta = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/form/input[3]'))
        )
        campo_coleta.clear()
        campo_coleta.send_keys(coleta)
        campo_coleta.send_keys(Keys.ENTER)
        print(f"🔍 Coleta {coleta} consultada")

        time.sleep(4)
        campo_coleta.send_keys(Keys.CONTROL, '103')
        time.sleep(5)
        contador = contador +1
        faltacoletas = qtdcoletas - contador
        print(f'Foram consultadas {contador} coletas. Faltam consultar {faltacoletas} coletas')

    print("✅ Todas as coletas visíveis do Excel foram consultadas.")
    
    # Abrir o Excel no final para colocar as justificativas
    os.startfile(arquivo_excel)
    input("🟢 Pressione Enter para encerrar...")

except Exception as e:
    print(f"❌ Erro: {e}")
input('nao fecha')