import os
import asyncio
from playwright.async_api import async_playwright
from datetime import datetime
import pandas as pd
import openpyxl

# --- CONFIGURAÇÕES DE ACESSO E CAMINHOS ---
USER_PBI = "danilo.lacerda@transking.com.br"
PASS_PBI = "QwErTyAsDfgH!2#4%6"

# LINK ATUALIZADO DO SEU BI
REPORT_URL = "https://app.powerbi.com/groups/59efcb19-4c50-47fc-8dd1-cdcee2357723/reports/e1a37785-9bef-43e0-8354-8392ecd3bf6f/eb432ee736f001c6abc0?experience=power-bi"

# PASTA DE DESTINO DO SEU GITHUB
PASTA_DESTINO = r"C:\Users\flavi\OneDrive - transking.com.br\PROCESSOS\01 - Equipe\04 - Weslley"

async def exportar_tabela_powerbi():
    pasta_do_script = os.path.dirname(os.path.abspath(__file__))
    
    async with async_playwright() as p:
        print(f"🚀 Iniciando navegador...")
        
        browser = await p.chromium.launch(
            headless=False,
            args=["--start-maximized", "--window-position=0,0"]
        )

        context = await browser.new_context(
            accept_downloads=True,
            viewport=None 
        )
        
        page = await context.new_page()

        print(f"🔗 Acessando Power BI...")
        await page.goto(REPORT_URL)

        try:
            # --- LOGIN MICROSOFT ---
            print("🔑 Preenchendo e-mail...")
            await page.wait_for_selector('#email', timeout=20000)
            await page.fill('#email', USER_PBI)
            await page.press('#email', 'Enter')

            print("🔑 Preenchendo senha...")
            await page.wait_for_selector('input[type="password"]', timeout=15000)
            await page.fill('input[type="password"]', PASS_PBI)
            await page.click('input[type="submit"]')

            try:
                await page.wait_for_selector('#idSIButton9', timeout=8000)
                await page.click('#idSIButton9')
                print("✅ Login efetuado com sucesso!")
            except:
                print("⚠️ Botão de 'Manter conectado' não apareceu, prosseguindo...")

            # --- INTERVALO PARA CARREGAMENTO ---
            print("⏳ Aguardando 5 segundos para carregar o Power BI...")
            await asyncio.sleep(5) 
            
            print(f"🔄 Recarregando a página do relatório...")
            await page.goto(REPORT_URL)
            
            print("⏳ Aguardando 5 segundos adicionais para renderização...")
            await asyncio.sleep(5)

            await page.bring_to_front()
            
            # --- FLUXO DE CLIQUES ---
            print("🔍 Procurando a coluna 'DESTINO' para ativação...")
            seletor_coluna_destino = 'div[role="columnheader"][data-query-ref="Min(Coletas.DESTINO)"]'
            coluna_destino = await page.wait_for_selector(seletor_coluna_destino, timeout=60000)
            
            print("Main-cell 🖱️ Passando o mouse na coluna 'DESTINO' para ativar o visual...")
            await coluna_destino.hover()
            await page.wait_for_timeout(1000)
            await coluna_destino.click()
            await page.wait_for_timeout(2000)

            print("🖱️ Clicando nos 3 pontinhos (...)")
            seletor_tres_pontos = 'button[data-testid="visual-more-options-btn"]'
            await page.wait_for_selector(seletor_tres_pontos, state="visible", timeout=20000)
            await page.click(seletor_tres_pontos)
            
            print("📂 Selecionando 'Exportar dados'...")
            seletor_menu_exportar = 'button[data-testid="pbimenu-item.Exportar dados"]'
            await page.wait_for_selector(seletor_menu_exportar, state="visible", timeout=15000)
            await page.click(seletor_menu_exportar)

            print("⏳ Aguardando janela de confirmação de exportação...")
            seletor_botao_confirmar = 'button[data-testid="export-btn"]'
            await page.wait_for_selector(seletor_botao_confirmar, state="visible", timeout=15000)

            print("🔘 Clicando em Exportar e baixando o arquivo...")
            async with page.expect_download() as download_info:
                await page.click(seletor_botao_confirmar)
            
            download = await download_info.value
            
            if not os.path.exists(PASTA_DESTINO):
                os.makedirs(PASTA_DESTINO)

            # 1. Salva o arquivo original bruto
            caminho_final = os.path.join(PASTA_DESTINO, "RelatorioAtualizado.xlsx")
            await download.save_as(caminho_final)
            print(f"📍 RELATÓRIO BRUTO SALVO EM: {caminho_final}")

            # --- TRATAMENTO DE DADOS PARA OS ANALISTAS ---
            print("⚙️ Tratando dados para o arquivo de Monitoramento...")
            
            df_novo = pd.read_excel(caminho_final)
            df_novo.columns = df_novo.columns.str.strip()

            # Força colunas numéricas de validação
            for col_int in ["COLETA", "COTAÇÃO"]:
                if col_int in df_novo.columns:
                    df_novo[col_int] = pd.to_numeric(df_novo[col_int], errors='coerce').fillna(0).astype(int)

            # Remove as linhas onde COTAÇÃO ou COLETA são iguais a 0
            if "COTAÇÃO" in df_novo.columns:
                df_novo = df_novo[df_novo["COTAÇÃO"] != 0]
            if "COLETA" in df_novo.columns:
                df_novo = df_novo[df_novo["COLETA"] != 0]

            # Filtrar os status indesejados
            status_remover = ["VALIDAÇÃO FINAL", "1. AG. PLANEJAMENTO"]
            if "STATUS" in df_novo.columns:
                df_novo = df_novo[~df_novo["STATUS"].astype(str).str.strip().isin(status_remover)]

            # Lista de colunas para remover
            colunas_remover = ["R$ MERCADORIA", "MOPP", "VEÍCULO", "CARRO", "CARGA DESCARGA"]
            df_novo = df_novo.drop(columns=[col for col in colunas_remover if col in df_novo.columns])

            # Tratamento de datas
            for col_data in ["PROG. COLETA", "PROG. ENTREGA"]:
                if col_data in df_novo.columns:
                    df_novo[col_data] = pd.to_datetime(df_novo[col_data], errors='coerce', format='mixed')

            caminho_monitoramento = os.path.join(PASTA_DESTINO, "AcompanhamentoMonitoramento.xlsx")

            # --- NOVO BLOCO DE HISTÓRICO (MULTÍPLAS ABAS) ---
            df_historico_consolidado = pd.DataFrame()
            
            if os.path.exists(caminho_monitoramento):
                print("🔄 Arquivo antigo de abas encontrado. Consolidando histórico das guias...")
                try:
                    # Carrega todas as abas existentes no arquivo atual dos analistas
                    abas_antigas = pd.read_excel(caminho_monitoramento, sheet_name=None)
                    lista_dfs = []
                    
                    for nome_aba, df_aba in abas_antigas.items():
                        df_aba.columns = df_aba.columns.str.strip()
                        # Garante que as chaves estejam tratadas uniformemente
                        for col_int in ["COLETA", "COTAÇÃO"]:
                            if col_int in df_aba.columns:
                                df_aba[col_int] = pd.to_numeric(df_aba[col_int], errors='coerce').fillna(0).astype(int)
                        
                        if "MONITORAMENTO" in df_aba.columns and "STATUS" in df_aba.columns:
                            lista_dfs.append(df_aba[["COTAÇÃO", "COLETA", "STATUS", "MONITORAMENTO"]].copy())
                    
                    if lista_dfs:
                        df_historico_consolidado = pd.concat(lista_dfs, ignore_index=True)
                        df_historico_consolidado = df_historico_consolidado.drop_duplicates(subset=["COTAÇÃO", "COLETA"])
                        df_historico_consolidado = df_historico_consolidado.rename(columns={"STATUS": "STATUS_ANTIGO"})
                except Exception as e:
                    print(f"⚠️ Alerta ao ler histórico das abas antigas: {e}")

            # Cruzamento do histórico com o df_novo
            if not df_historico_consolidado.empty:
                df_novo = pd.merge(df_novo, df_historico_consolidado, on=["COTAÇÃO", "COLETA"], how="left")
                
                # Regra do status: preserva o texto se o status atual for igual ao status antigo
                df_novo["MONITORAMENTO"] = df_novo.apply(
                    lambda row: row["MONITORAMENTO"] if str(row["STATUS"]).strip() == str(row["STATUS_ANTIGO"]).strip() else None,
                    axis=1
                )
                df_novo = df_novo.drop(columns=["STATUS_ANTIGO"])
            else:
                if "MONITORAMENTO" not in df_novo.columns:
                    df_novo["MONITORAMENTO"] = None

            # Tratamento da coluna CURVA para evitar erros na criação das guias
            if "CURVA" in df_novo.columns:
                df_novo["CURVA"] = df_novo["CURVA"].astype(str).str.strip().replace({'nan': 'SEM CURVA', '': 'SEM CURVA'})
                # Obtém todas as curvas únicas presentes na base
                curvas_unicas = df_novo["CURVA"].unique()
            else:
                df_novo["CURVA"] = "GERAL"
                curvas_unicas = ["GERAL"]

            # --- GRAVAÇÃO EM MÚLTIPLAS ABAS FILTRADAS E ORDENADAS POR COTAÇÃO ---
            print("💾 Separando dados por CURVA e ordenando por COTAÇÃO...")
            with pd.ExcelWriter(caminho_monitoramento, engine='openpyxl') as writer:
                for curva in sorted(curvas_unicas):
                    df_curva = df_novo[df_novo["CURVA"] == curva].copy()
                    
                    # 🔥 CLASSIFICAÇÃO DA BASE PELA COLUNA COTAÇÃO (Menor para Maior)
                    if "COTAÇÃO" in df_curva.columns:
                        df_curva = df_curva.sort_values(by="COTAÇÃO", ascending=True)
                        
                    # Cria uma aba para cada curva (o Excel limita nomes de abas a 31 caracteres)
                    nome_aba = curva[:31]
                    df_curva.to_excel(writer, sheet_name=nome_aba, index=False)

            # --- PROCESSAMENTO EXCEL COM OPENPYXL PARA FORMATAÇÃO (TODAS AS ABAS) ---
            print("🎨 Aplicando formatação visual nas abas...")
            wb = openpyxl.load_workbook(caminho_monitoramento)
            
            for ws in wb.worksheets:
                colunas_datas = []
                coluna_cpf_idx = None

                for col_idx in range(1, ws.max_column + 1):
                    nome_coluna = ws.cell(row=1, column=col_idx).value
                    if nome_coluna in ["PROG. COLETA", "PROG. ENTREGA"]:
                        colunas_datas.append(col_idx)
                    elif nome_coluna == "CPF":
                        coluna_cpf_idx = col_idx

                # Formatação de datas
                for col_idx in colunas_datas:
                    for row_idx in range(2, ws.max_row + 1):
                        celula = ws.cell(row=row_idx, column=col_idx)
                        if celula.value is not None:
                            celula.number_format = 'dd/mm/yyyy hh:mm'

                # Formatação de CPF
                if coluna_cpf_idx:
                    for row_idx in range(2, ws.max_row + 1):
                        celula = ws.cell(row=row_idx, column=coluna_cpf_idx)
                        if celula.value is not None:
                            celula.number_format = '0'

            wb.save(caminho_monitoramento)
            wb.close()

            print("\n" + "="*60)
            print(f"🎉 PROCESSO CONCLUÍDO COM SUCESSO!")
            print(f"📍 RELATÓRIO BRUTO: {caminho_final}")
            print(f"📍 MONITORAMENTO POR ABAS: {caminho_monitoramento}")
            print("="*60 + "\n")

        except Exception as e:
            print(f"❌ Erro durante o processo: {e}")
            try:
                caminho_erro = os.path.join(pasta_do_script, "erro_captura.png")
                await page.screenshot(path=caminho_erro)
                print(f"📸 Evidência de erro gravada em: {caminho_erro}")
            except:
                pass
        
        finally:
            print("🧹 Encerrando a instância do navegador...")
            await page.wait_for_timeout(4000)
            await browser.close()

if __name__ == "__main__":
    asyncio.run(exportar_tabela_powerbi())