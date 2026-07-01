import os
import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import openpyxl

# --- CONFIGURAÇÕES DE ACESSO E CAMINHOS ---
USER_PBI = "danilo.lacerda@transking.com.br"
PASS_PBI = "QwErTyAsDfgH!2#4%6"

# LINK DO SEU BI
REPORT_URL = "https://app.powerbi.com/groups/59efcb19-4c50-47fc-8dd1-cdcee2357723/reports/e1a37785-9bef-43e0-8354-8392ecd3bf6f/eb432ee736f001c6abc0?experience=power-bi"

# PASTA DE DESTINO DO ONEDRIVE
PASTA_DESTINO = r"C:\Users\flavi\OneDrive - transking.com.br\PROCESSOS\01 - Equipe\04 - Weslley"

async def exportar_tabela_powerbi():
    pasta_do_script = os.path.dirname(os.path.abspath(__file__))
    
    async with async_playwright() as p:
        print(f"🚀 Iniciando navegador...")
        
        browser = await p.chromium.launch(
            headless=False,
            args=["--start-maximized", "--window-position=0,0"]
        )

        context = await browser.new_context(
            accept_downloads=True,
            viewport=None 
        )
        
        page = await context.new_page()

        print(f"🔗 Acessando Power BI...")
        await page.goto(REPORT_URL)

        try:
            # --- LOGIN MICROSOFT ---
            print("🔑 Preenchendo e-mail...")
            await page.wait_for_selector('#email', timeout=20000)
            await page.fill('#email', USER_PBI)
            await page.press('#email', 'Enter')

            print("🔑 Preenchendo senha...")
            await page.wait_for_selector('input[type="password"]', timeout=15000)
            await page.fill('input[type="password"]', PASS_PBI)
            await page.click('input[type="submit"]')

            try:
                await page.wait_for_selector('#idSIButton9', timeout=8000)
                await page.click('#idSIButton9')
                print("✅ Login efetuado com sucesso!")
            except:
                print("⚠️ Botão de 'Manter conectado' não apareceu, prosseguindo...")

            # --- INTERVALO PARA CARREGAMENTO ---
            print("⏳ Aguardando 5 segundos para carregar o Power BI...")
            await asyncio.sleep(5) 
            
            print(f"🔄 Recarregando a página do relatório...")
            await page.goto(REPORT_URL)
            
            print("⏳ Aguardando 5 segundos adicionais para renderização...")
            await asyncio.sleep(5)

            await page.bring_to_front()
            
            # --- FLUXO DE CLIQUES ---
            print("🔍 Procurando a coluna 'DESTINO' para ativação...")
            seletor_coluna_destino = 'div[role="columnheader"][data-query-ref="Min(Coletas.DESTINO)"]'
            coluna_destino = await page.wait_for_selector(seletor_coluna_destino, timeout=60000)
            
            print("Main-cell 🖱️ Passando o mouse na coluna 'DESTINO' para ativar o visual...")
            await coluna_destino.hover()
            await page.wait_for_timeout(1000)
            await coluna_destino.click()
            await page.wait_for_timeout(2000)

            print("🖱️ Clicando nos 3 pontinhos (...)")
            seletor_tres_pontos = 'button[data-testid="visual-more-options-btn"]'
            await page.wait_for_selector(seletor_tres_pontos, state="visible", timeout=20000)
            await page.click(seletor_tres_pontos)
            
            print("📂 Selecionando 'Exportar dados'...")
            seletor_menu_exportar = 'button[data-testid="pbimenu-item.Exportar dados"]'
            await page.wait_for_selector(seletor_menu_exportar, state="visible", timeout=15000)
            await page.click(seletor_menu_exportar)

            print("⏳ Aguardando janela de confirmação de exportação...")
            seletor_botao_confirmar = 'button[data-testid="export-btn"]'
            await page.wait_for_selector(seletor_botao_confirmar, state="visible", timeout=15000)

            print("🔘 Clicando em Exportar e baixando o arquivo...")
            async with page.expect_download() as download_info:
                await page.click(seletor_botao_confirmar)
            
            download = await download_info.value
            
            if not os.path.exists(PASTA_DESTINO):
                os.makedirs(PASTA_DESTINO)

            # Caminho temporário para salvar o arquivo bruto baixado do Power BI
            caminho_temporario = os.path.join(PASTA_DESTINO, "temp_pbi_download.xlsx")
            await download.save_as(caminho_temporario)

            # --- TRATAMENTO DOS DADOS NO PANDAS ---
            print("⚙️ Iniciando o tratamento dos dados...")
            df = pd.read_excel(caminho_temporario)
            
            # Limpa espaços em branco nos nomes das colunas
            df.columns = df.columns.str.strip()

            # 1. FILTRO: Coluna CLIENTE deve conter "UNILEVER"
            if "CLIENTE" in df.columns:
                df = df[df["CLIENTE"].astype(str).str.upper().str.contains("UNILEVER", na=False)]

            # 2. FILTRO: Coluna STATUS aceita apenas os 3 status específicos
            status_permitidos = ["1. AG. PLANEJAMENTO", "2. EM GR", "3. AG. COLETA"]
            if "STATUS" in df.columns:
                df = df[df["STATUS"].astype(str).str.strip().isin(status_permitidos)]

            # 3. REMOÇÃO DE COLUNAS CONFORME O SEU PADRÃO
            colunas_remover = ["R$ MERCADORIA", "MOPP", "VEÍCULO", "CARRO", "CARGA DESCARGA"]
            df = df.drop(columns=[col for col in colunas_remover if col in df.columns])

            # 4. TRATAMENTO DA COLUNA ID: Trazer apenas os últimos 8 caracteres
            if "ID" in df.columns:
                df["ID"] = df["ID"].astype(str).str.strip().str[-8:]

            # 5. REMOÇÃO DE DUPLICADAS NA COLUNA COTAÇÃO:
            # Mantém apenas a primeira ocorrência se estiver duplicada. Se estiver em branco, ignora.
            if "COTAÇÃO" in df.columns:
                mascara_manter = ~df.duplicated(subset=["COTAÇÃO"], keep="first") | df["COTAÇÃO"].isna() | (df["COTAÇÃO"].astype(str).str.strip() == "")
                df = df[mascara_manter]

            # Garante que o Pandas entenda as colunas de data para a formatação posterior
            for col_data in ["PROG. COLETA", "PROG. ENTREGA"]:
                if col_data in df.columns:
                    df[col_data] = pd.to_datetime(df[col_data], errors='coerce', format='mixed')

            # Caminho final do arquivo unificado
            caminho_final = os.path.join(PASTA_DESTINO, "Unilevercancelamentos.xlsx")
            df.to_excel(caminho_final, index=False)

            # --- CORREÇÃO DE FORMATAÇÃO VISUAL DIRETAMENTE NO EXCEL ---
            wb = openpyxl.load_workbook(caminho_final)
            ws = wb.active

            colunas_datas = []
            coluna_cpf_idx = None

            for col_idx in range(1, ws.max_column + 1):
                nome_coluna = ws.cell(row=1, column=col_idx).value
                if nome_coluna in ["PROG. COLETA", "PROG. ENTREGA"]:
                    colunas_datas.append(col_idx)
                elif nome_coluna == "CPF":
                    coluna_cpf_idx = col_idx

            # Aplica máscara visual nas datas
            for col_idx in colunas_datas:
                for row_idx in range(2, ws.max_row + 1):
                    celula = ws.cell(row=row_idx, column=col_idx)
                    if celula.value is not None:
                        celula.number_format = 'dd/mm/yyyy hh:mm'

            # Corrigindo o CPF para não exibir notação científica
            if coluna_cpf_idx:
                for row_idx in range(2, ws.max_row + 1):
                    celula = ws.cell(row=row_idx, column=coluna_cpf_idx)
                    if celula.value is not None:
                        celula.number_format = '0'

            wb.save(caminho_final)
            wb.close()

            # Apaga o arquivo temporário bruto para deixar a pasta limpa
            if os.path.exists(caminio_temporario := caminho_temporario):
                os.remove(caminho_temporario)

            print("\n" + "="*60)
            print(f"🎉 PROCESSO CONCLUÍDO COM SUCESSO!")
            print(f"📍 ARQUIVO GERADO: {caminho_final}")
            print("="*60 + "\n")

        except Exception as e:
            print(f"❌ Erro durante o processo: {e}")
            try:
                caminho_erro = os.path.join(pasta_do_script, "erro_captura.png")
                await page.screenshot(path=caminho_erro)
                print(f"📸 Evidência de erro gravada em: {caminho_erro}")
            except:
                pass
        
        finally:
            print("🧹 Encerrando a instância do navegador...")
            await page.wait_for_timeout(4000)
            await browser.close()

if __name__ == "__main__":
    asyncio.run(exportar_tabela_powerbi())


