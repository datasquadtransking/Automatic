import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Caminho do arquivo original
caminho_arquivo = r"C:\Users\flavi\Desktop\UnileverPontualidade.xlsx"

# Leitura da planilha
df = pd.read_excel(caminho_arquivo)

# Seleção das colunas relevantes
colunas_desejadas = [
    "Shipment",
    "SAP ID",
    "Origem",
    "Dock",
    "CH",
    "Pontualidade",
    "Chegada BRLOG",
    "Pontualidade BRLOG"
]
df_filtrado = df[colunas_desejadas].copy()

# --- 🔹 Formatação de colunas de data/hora ---
for coluna in df_filtrado.columns:
    if df_filtrado[coluna].dtype == 'datetime64[ns]' or df_filtrado[coluna].astype(str).str.contains(r'\d{4}-\d{2}-\d{2}').any():
        df_filtrado[coluna] = pd.to_datetime(df_filtrado[coluna], errors='coerce')
        # Se há hora válida, mostra hora também
        if df_filtrado[coluna].dt.time.nunique() > 1:
            df_filtrado[coluna] = df_filtrado[coluna].dt.strftime("%d/%m/%y %H:%M")
        else:
            df_filtrado[coluna] = df_filtrado[coluna].dt.strftime("%d/%m/%y")

# Cria coluna Data com base na coluna Dock
df_filtrado["Data"] = pd.to_datetime(df["Dock"], errors="coerce").dt.date

# Remove linhas sem data válida
df_filtrado = df_filtrado.dropna(subset=["Data"])

# Formata coluna Data para dd/mm/aa na base original
df_filtrado["Data"] = pd.to_datetime(df_filtrado["Data"]).dt.strftime("%d/%m/%y")

# --- 🔹 Agrupamento diário ---
tabela_detalhada = df_filtrado.groupby("Data").agg(
    No_Prazo=("Pontualidade", lambda x: (x == "ON TIME").sum()),
    Fora_do_Prazo=("Pontualidade", lambda x: (x == "NO SHOW").sum()),
    Total=("Pontualidade", "count")
).reset_index()

# Calcula SLA diário
tabela_detalhada["SLA"] = tabela_detalhada["No_Prazo"] / tabela_detalhada["Total"]

# Adiciona coluna de Contestação
tabela_detalhada["Contestação"] = 0

# Calcula SLA acumulado
tabela_detalhada["SLA Acumulado"] = (
    tabela_detalhada["No_Prazo"].cumsum() / tabela_detalhada["Total"].cumsum()
)

# Formata a data
tabela_detalhada["Data"] = pd.to_datetime(tabela_detalhada["Data"]).dt.strftime("%d/%m/%y")

# --- 🔹 Adiciona linha de total ---
linha_total = pd.DataFrame({
    "Data": ["Total"],
    "No_Prazo": [tabela_detalhada["No_Prazo"].sum()],
    "Fora_do_Prazo": [tabela_detalhada["Fora_do_Prazo"].sum()],
    "Total": [tabela_detalhada["Total"].sum()],
    "SLA": [tabela_detalhada["No_Prazo"].sum() / tabela_detalhada["Total"].sum()],
    "Contestação": [0],
    "SLA Acumulado": [None]
})

# Concatena a linha de total ao final
tabela_detalhada = pd.concat([tabela_detalhada, linha_total], ignore_index=True)

# Formata percentuais
tabela_detalhada["SLA"] = tabela_detalhada["SLA"].apply(lambda x: f"{x*100:.2f}%" if pd.notnull(x) else "")
tabela_detalhada["SLA Acumulado"] = tabela_detalhada["SLA Acumulado"].apply(lambda x: f"{x*100:.2f}%" if pd.notnull(x) else "")

# --- 🔹 Exporta para Excel ---
caminho_saida = r"C:\Users\flavi\Desktop\UnileverPontualidade_Com_Detalhamento.xlsx"

with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
    df_filtrado.to_excel(writer, sheet_name="Base Original", index=False)
    tabela_detalhada.to_excel(writer, sheet_name="Resumo por Dia", index=False)

# --- 🔹 Formatação da linha total no Excel ---
wb = load_workbook(caminho_saida)
ws = wb["Resumo por Dia"]

ultima_linha = ws.max_row

# Estilos: negrito + fundo cinza
fonte_negrito = Font(bold=True)
preenchimento_cinza = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Aplica estilo à última linha
for celula in ws[ultima_linha]:
    celula.font = fonte_negrito
    celula.fill = preenchimento_cinza

wb.save(caminho_saida)

print("✅ Arquivo exportado com datas formatadas, linha total somada e destacada no resumo!")
